from fastapi import FastAPI, HTTPException
from fastapi.openapi.docs import get_swagger_ui_html, get_redoc_html
from fastapi.staticfiles import StaticFiles
from openpyxl import load_workbook
from pydantic import BaseModel
from typing import List, Optional

app = FastAPI()

class Escuela(BaseModel):
    Clave_del_centro_de_trabajo: str
    Clave_del_turno: int
    Nombre_del_turno: str
    Nombre_del_centro_de_trabajo: str
    Tipo_educativo: str
    Nivel_educativo: str
    Servicio_educativo: str
    Nombre_del_control: str
    Tipo_de_sostenimiento: str
    Clave_de_la_entidad_federativa: int
    Nombre_de_la_entidad: str
    Clave_del_municipio_o_delegacion: int
    Nombre_del_municipio_o_delegacion: str
    Clave_de_la_localidad: int
    Nombre_de_localidad: str
    Domicilio: str
    Numero_exterior: int
    Entre_la_calle: str
    Y_la_calle: str
    Calle_posterior: str
    Colonia: int
    Nombre_de_la_colonia: str
    Codigo_postal: int
    Clave_lada: Optional[int]
    Telefono: Optional[int]
    Extension_del_telefono: Optional[int] = None
    Pagina_web: Optional[str] = None
    Alumnos_total_hombres: Optional[int]
    Alumnos_total_mujeres: Optional[int]
    Alumnos_total: Optional[int]
    Docentes_total_hombres: Optional[int]
    Docentes_total_mujeres: Optional[int]
    Docentes_total: Optional[int]
    Aulas_en_uso: Optional[int]
    Aulas_existentes: Optional[int]
    Tipo_de_localidad: Optional[str] = None
    Ubicacion_escuela_localidad_al_oeste_del_meridiano_de_greenwich_expresada_en_grados_minutos_y_segundos: str
    Ubicacion_escuela_localidad_al_norte_del_ecuador_expresada_en_grados_minutos_y_segundos: str
    Ubicacion_escuela_localidad_al_oeste_del_meridiano_de_greenwich_expresada_en_grados: float
    Ubicacion_escuela_localidad_al_norte_del_ecuador_expresada_en_grados: float

@app.get("/api/escuelas/{codigo_postal}", response_model=List[Escuela])
async def search_excel(codigo_postal: str):
    """
    Busca una escuela por código postal en el archivo "schools.xlsx" y devuelve una lista con los datos de las escuelas encontradas como JSON.

    Args:
      codigo_postal: Código postal de las escuelas a buscar.

    Returns:
      Lista con los datos de las escuelas o un mensaje de error si no se encuentran.
    """

    workbook = load_workbook("schools.xlsx")
    worksheet = workbook.active

    escuelas = []

    for row in worksheet.iter_rows():
        if str(row[22].value) == codigo_postal:
            data = {}
            for i, cell in enumerate(row):
                data[worksheet.cell(row=1, column=i + 1).value] = cell.value
            escuela = Escuela(**data)
            escuelas.append(escuela)

    if escuelas:
        return escuelas
    raise HTTPException(status_code=404, detail="No se encontraron escuelas con el código postal especificado")

# Rutas para la documentación
app.mount("/docs", StaticFiles(directory="docs"), name="docs")
app.get("/docs", include_in_schema=False)(get_swagger_ui_html)
app.get("/redoc", include_in_schema=False)(get_redoc_html)


